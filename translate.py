import requests
import hashlib
import random
from openpyxl import load_workbook
import argparse
import sqlite3
import os
from langdetect import detect, DetectorFactory, LangDetectException
import re
import unicodedata
from dotenv import load_dotenv

load_dotenv()
DetectorFactory.seed = 0

# 百度翻译配置
APP_ID = os.getenv("BAIDU_APP_ID")
APP_KEY = os.getenv("BAIDU_APP_KEY")
TRANSLATE_API = 'https://fanyi-api.baidu.com/api/trans/vip/translate'
LANGUAGE_CODE = [
    # 自动检测
    'auto',
    # 中文
    'zh',
    # 英语
    'en',
    # 粤语
    'yue',
    # 文言文
    'wyw',
    # 日语,
    'jp',
    # 韩语
    'kor',
    # 法语
    'fra',
    # 西班牙语
    'spa',
    # 泰语
    'th',
    # 阿拉伯语
    'ara',
    # 俄语
    'ru',
    # 葡萄牙语
    'pt',
    # 德语
    'de',
    # 意大利语
    'it',
    # 希腊语
    'el',
    # 荷兰语
    'nl',
    # 波兰语
    'pl',
    # 保加利亚语
    'bul',
    # 爱沙尼亚语
    'est',
    # 丹麦语
    'dan',
    # 芬兰语
    'fin',
    # 捷克语
    'cs',
    # 罗马尼亚语
    'rom',
    # 斯洛文尼亚语
    'slo',
    # 瑞典语
    'swe',
    # 匈牙利语
    'hu',
    # 繁体中文
    'cht',
    # 越南语
    'vie',
]


# 如果是纯数字,则不翻译
def is_number(s):
    try:
        float(s)
        return True
    except (TypeError, ValueError):
        pass

    try:
        unicodedata.numeric(s)
        return True
    except (TypeError, ValueError):
        pass

    return False


# 如果不包含源语言,则不翻译
def is_need_translate(text, from_lang):
    lang_map = {
        "af": "auto",
        "ar": "ara",
        "bg": "bul",
        "cs": "cs",
        "da": "dan",
        "de": "de",
        "el": "el",
        "en": "en",
        "es": "spa",
        "et": "est",
        "fa": "fa",
        "fi": "fin",
        "fr": "fra",
        "hi": "hi",
        "hr": "hr",
        "hu": "hu",
        "id": "id",
        "it": "it",
        "ja": "jp",
        "ko": "kor",
        "lt": "lt",
        "lv": "lv",
        "pl": "pl",
        "pt": "pt",
        "ro": "rom",
        "ru": "ru",
        "sl": "slo",
        "sv": "swe",
        "sw": "sw",
        "ta": "ta",
        "th": "th",
        "tr": "tr",
        "uk": "uk",
        "ur": "ur",
        "vi": "vie",
        "zh-cn": "zh",
        "zh-tw": "cht",
    }
    try:
        language = detect(text)
    except LangDetectException:
        return False
    if language not in lang_map:
        return False
    return lang_map[language] == from_lang


# 如果是公式则不翻译
def is_formula(text):
    try:
        # 匹配常见公式符号的正则表达式
        formula_pattern = r"[=]"
        return bool(re.search(formula_pattern, text))
    except (TypeError, ValueError):
        return False


# 调用百度翻译
def baidu_translate(text, from_lang, to_lang):
    salt = str(random.randint(327681365, 655368412))

    def translate_line(line):
        sign = hashlib.md5(
            (APP_ID + line + salt + APP_KEY).encode('utf-8')
        ).hexdigest()
        params = {
            'q': line,
            'from': from_lang,
            'to': to_lang,
            'appid': APP_ID,
            'salt': salt,
            'sign': sign,
        }
        response = requests.get(TRANSLATE_API, params=params).json()
        return response.get('trans_result', [{}])[0].get('dst', line)

    # 按行翻译，并保持换行
    lines = text.splitlines()
    translated_lines = [translate_line(line) for line in lines]
    return '\n'.join(translated_lines)


# 从数据库中获取 key
def get_from_local_db(conn, key, to_lang):
    cursor = conn.cursor()
    cursor.execute(
        "SELECT value FROM kv_store WHERE key = ? AND to_lang = ? ",
        (
            key,
            to_lang,
        ),
    )
    result = cursor.fetchone()
    return result[0] if result else None


# 写入数据库
def write_to_local_db(conn, key, to_lang, value):
    cursor = conn.cursor()
    cursor.execute(
        "INSERT OR REPLACE INTO kv_store (key, to_lang, value) VALUES (?, ?, ?)",
        (key, to_lang, value),
    )
    conn.commit()


def to_translate(key, from_lang, to_lang, conn):
    # if is_need_translate(key, from_lang):
    value = get_from_local_db(conn, key, to_lang)
    if value is None:
        value = baidu_translate(key, from_lang, to_lang)
        if value == key:
            print(f"百度翻译出错,请检查key. text:{key} value:{value}")
            # exit()
        else:
            print(f'form {key} to {value}')
        write_to_local_db(conn, key, to_lang, value)
        return value
    return value


def start_translate(in_dir, out_dir, file_name, from_lang, to_lang, conn):
    # 打开 Excel 文件
    file_path = os.path.join(in_dir, file_name)
    workbook = load_workbook(file_path)
    print(f"正在翻译Excel: {file_name}")

    # 遍历所有工作表
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"正在翻译sheet: {sheet_name}")
        new_sheet_name = to_translate(sheet_name, from_lang, to_lang, conn)
        sheet.title = new_sheet_name.replace("/", "_").replace(":", "_")

        # 遍历每个单元格并翻译文本
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        ):
            for cell in row:
                if cell.value is None:
                    continue
                if is_number(cell.value):
                    continue
                if is_formula(cell.value):
                    continue
                # 仅翻译文本类型的单元格
                if isinstance(cell.value, str):
                    if len(cell.value.strip()) == 0:
                        continue
                    value = to_translate(cell.value, from_lang, to_lang, conn)
                    cell.value = value

    out_file_name = f'{to_lang}_{file_name}'
    out_file_path = os.path.join(out_dir, out_file_name)
    workbook.save(out_file_path)
    print("翻译完成")


# 初始化 SQLite 数据库
def init_db():
    db_file = "local_db.sqlite"
    is_new_db = not os.path.exists(db_file)
    conn = sqlite3.connect(db_file)

    # 如果是新建数据库，初始化表结构
    if is_new_db:
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE "main"."kv_store" (
                "key" TEXT NOT NULL,
                "value" TEXT,
                "to_lang" TEXT NOT NULL,
                PRIMARY KEY ("key", "to_lang")
            )
        """
        )
        conn.commit()
    return conn


def is_excel_file(file_path):
    return file_path.lower().endswith(('.xls', '.xlsx'))


def main(args):
    # 判断输入的参数是否正确
    if not os.path.exists(args.in_dir):
        raise FileNotFoundError(f"输入目录 {args.in_dir} 不存在！")
    if not os.path.exists(args.out_dir):
        os.makedirs(args.out_dir)
    if args.from_lang not in LANGUAGE_CODE:
        raise ValueError(
            f"语言代码 {args.from_lang} 不在支持的列表 {LANGUAGE_CODE} 中！"
        )
    if args.to_lang not in LANGUAGE_CODE:
        raise ValueError(
            f"语言代码 {args.from_lang} 不在支持的列表 {LANGUAGE_CODE} 中！"
        )

    conn = init_db()

    # 只处理指定目录下的文件,不处理子目录下的文件
    with os.scandir(args.in_dir) as entries:
        files = [entry.name for entry in entries if entry.is_file()]

    for file_name in files:
        if is_excel_file(file_name):
            start_translate(
                args.in_dir,
                args.out_dir,
                file_name,
                args.from_lang,
                args.to_lang,
                conn,
            )


'''
eg.
python translate.py \
    --in_dir=./in/ \
    --out_dir=./out/ \
    --from_lang=jp \
    --to_lang=zh
'''
if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="将excel文件翻译为指定语言")
    parser.add_argument(
        '--in_dir', default='./test/', help='翻译目录 (default: ./)'
    )
    parser.add_argument(
        '--out_dir',
        default='./out/',
        help='输出目录 (default: ./out/)',
    )
    parser.add_argument(
        '--from_lang', default='jp', help='源语言 (default: jp)'
    )
    parser.add_argument(
        '--to_lang', default='en', help='目标语言 (default: zh)'
    )

    args = parser.parse_args()
    main(args)
